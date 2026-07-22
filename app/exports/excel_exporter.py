import os
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


def filling_excel(file_name, sheet_name, user_id, user_data, auto_data):
    user_name = user_data[user_id]['user_name']
    user_surname = user_data[user_id]['user_surname']
    user_email = user_data[user_id]['user_email']
    hash_password = user_data[user_id]['hash_password']
    auto_brand = auto_data['auto_brand']
    auto_model = auto_data['auto_model']
    auto_year = auto_data['auto_year']
    auto_mileage = auto_data['auto_mileage']

    if not file_name.endswith(".xlsx"):
        file_name = f"{file_name}.xlsx"

    file_exists = os.path.exists(file_name)

    if file_exists:
        wb = load_workbook(file_name)
        sheet = wb.active
    else:
        wb = Workbook()
        sheet = wb.active
        if sheet_name:
            sheet.title = sheet_name
        _add_headers_and_styles(sheet)
        wb.save(file_name)

    last_row = sheet.max_row
    if last_row == 1 and sheet['A1'].value == "Имя":
        target_row = 2
    else:
        target_row = last_row + 1

    sheet[f'A{target_row}'] = user_name
    sheet[f'B{target_row}'] = user_surname
    sheet[f'C{target_row}'] = user_email
    sheet[f'D{target_row}'] = hash_password
    sheet[f'E{target_row}'] = auto_brand
    sheet[f'F{target_row}'] = auto_model
    sheet[f'G{target_row}'] = auto_year
    sheet[f'H{target_row}'] = auto_mileage
    wb.save(file_name)


def _add_headers_and_styles(sheet):
    sheet['A1'] = "Имя"
    sheet['B1'] = "Фамилия"
    sheet['C1'] = "Email"
    sheet['D1'] = "Пароль"
    sheet['E1'] = "Марка автомобиля"
    sheet['F1'] = "Модель автомобиля"
    sheet['G1'] = "Год выпуска"
    sheet['H1'] = "Пробег"

    sheet.column_dimensions["A"].width = 15
    sheet.column_dimensions["B"].width = 15
    sheet.column_dimensions["C"].width = 25
    sheet.column_dimensions["D"].width = 50
    sheet.column_dimensions["E"].width = 20
    sheet.column_dimensions["F"].width = 20
    sheet.column_dimensions["G"].width = 20
    sheet.column_dimensions["H"].width = 15

    for cell in "ABCDEFGH":
        header_cell = sheet[f"{cell}1"]
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(horizontal="center", vertical="center")

